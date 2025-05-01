import os
import sys
import streamlit as st
from typing import Optional, Tuple, List, Dict, Any
import mimetypes
import tempfile
import uuid
import pandas as pd
from io import BytesIO
from datetime import datetime

# Add parent directory to path to allow importing when run directly
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

# Import from parent modules
from OnPrem import read_config
from utils import setup_llm, load_llm, construct_link, check_create_symlink
from utils import lucene_to_chroma, get_prompt_template
from onprem.ingest.stores.sparse import SparseStore
from onprem.ingest.stores.dense import DenseStore
from onprem.ingest.stores.dual import DualStore


def sanitize_for_excel(text):
    """Helper function to sanitize text for Excel"""
    if isinstance(text, str):
        # Remove control characters and other characters that Excel can't handle
        chars_to_remove = '\x00-\x08\x0B-\x0C\x0E-\x1F\x7F'
        import re
        return re.sub(f'[{chars_to_remove}]', '', text)
    return text


def format_excel_writer(writer, df):
    """Apply Excel formatting for better appearance"""
    # Get the workbook and sheet
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # Create a bold format for headers
    from openpyxl.styles import Font, Alignment, PatternFill
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Apply formatting to headers
    for col_num, column_title in enumerate(df.columns):
        cell = worksheet.cell(row=1, column=col_num+1)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
    
    # Auto-adjust columns' width to fit content
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        # Add some padding (8 characters)
        adjusted_width = (length + 8)
        # Limit maximum width to avoid extremely wide columns
        column_width = min(adjusted_width, 100)
        worksheet.column_dimensions[column_cells[0].column_letter].width = column_width


def resolve_path(path):
    """
    Resolves any format strings in the given path
    
    Args:
        path: The path that might contain format strings
        
    Returns:
        The resolved path
    """
    if not path:
        return path
        
    # Import utils
    from onprem import utils as U
    
    # Handle common format strings
    resolved = path
    if "{webapp_dir}" in path:
        resolved = resolved.replace("{webapp_dir}", U.get_webapp_dir())
    if "{models_dir}" in path:
        resolved = resolved.replace("{models_dir}", U.get_models_dir())
    if "{datadir}" in path:
        resolved = resolved.replace("{datadir}", U.get_datadir())
        
    return resolved


def prepare_excel_file(results_df):
    """Create a formatted Excel file in memory"""
    # Apply sanitization to all text columns
    for col in results_df.columns:
        results_df[col] = results_df[col].apply(sanitize_for_excel)
    
    # Create Excel file in memory with enhanced formatting
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        results_df.to_excel(writer, index=False)
        # Apply formatting to make the Excel file more readable
        format_excel_writer(writer, results_df)
    output.seek(0)
    
    return output


def main():
    """
    Page for analyzing documents with a custom prompt
    """
    # Handle reset before any widgets are rendered
    if "reset_triggered" in st.session_state and st.session_state.reset_triggered:
        # Remove all session state except special keys
        keys_to_keep = []
        for key in list(st.session_state.keys()):
            if key not in keys_to_keep:
                del st.session_state[key]
        st.rerun()
    
    cfg = read_config()[0]
    
    # Get configuration
    RAG_SOURCE_PATH = cfg.get("ui", {}).get("rag_source_path", None)
    RAG_BASE_URL = cfg.get("ui", {}).get("rag_base_url", None)
    VECTORDB_PATH = cfg.get("llm", {}).get("vectordb_path", None)
    STORE_TYPE = cfg.get("llm", {}).get("store_type", "dense")
    
    # Resolve paths - we need the correct documents path
    # The config may use format strings, so directly resolve them
    from onprem import utils as U
    
    # Get the raw path from config and resolve any format strings
    DOCUMENTS_PATH = RAG_SOURCE_PATH
    if RAG_SOURCE_PATH and "{webapp_dir}" in RAG_SOURCE_PATH:
        DOCUMENTS_PATH = RAG_SOURCE_PATH.replace("{webapp_dir}", U.get_webapp_dir())
    
    # Load LLM to get model name
    llm = load_llm()
    MODEL_NAME = llm.model_name
    
    # This function may change RAG_SOURCE_PATH, but we'll keep our resolved DOCUMENTS_PATH
    RAG_SOURCE_PATH, RAG_BASE_URL = check_create_symlink(RAG_SOURCE_PATH, RAG_BASE_URL)
    
    # Add some CSS for better styling - same as in other pages
    st.markdown("""
    <style>
    /* Improve chat message styling */
    .stChatMessage {
        padding: 0.75rem;
        border-radius: 0.5rem;
        margin-bottom: 1rem;
        box-shadow: 0 1px 2px rgba(0,0,0,0.1);
    }
    
    /* Style the chat input */
    .stChatInputContainer {
        border-top: 1px solid #e0e0e0;
        padding-top: 1rem;
    }
    
    /* Make user messages stand out */
    .stChatMessage[data-testid="stChatMessage-user"] {
        background-color: #f0f7ff;
    }
    
    /* Style the assistant messages */
    .stChatMessage[data-testid="stChatMessage-assistant"] {
        background-color: #f9f9f9;
    }
    
    /* System messages should be subtle */
    .stChatMessage[data-testid="stChatMessage-system"] {
        background-color: #f0f0f0;
        font-style: italic;
    }
    </style>
    """, unsafe_allow_html=True)

    # Improved page header - same style as in other pages
    st.markdown("""
    <h1 style="
        color: #0068c9;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #0068c9;
        margin-bottom: 1.5rem;
        display: flex;
        align-items: center;
    ">
        <span style="font-size: 1.8rem; margin-right: 0.5rem;">📊</span> 
        Document Analysis
    </h1>
    """, unsafe_allow_html=True)
    
    # Enhanced info card - similar to other pages
    st.markdown(f"""
    <div style="
        padding: 12px 15px;
        border-radius: 8px;
        margin-bottom: 25px;
        background: linear-gradient(to right, rgba(0, 104, 201, 0.05), rgba(92, 137, 229, 0.1));
        border-left: 4px solid #0068c9;
        display: flex;
        align-items: center;
    ">
        <span style="font-size: 1.2rem; margin-right: 10px;">📊</span>
        <div>
            <p style="margin: 0; font-weight: 500;">Current Model: <span style="color: #0068c9;">{MODEL_NAME}</span></p>
            <p style="margin: 3px 0 0 0; font-size: 0.85rem; color: #666;">
                Apply a custom prompt to document chunks and export results to Excel for further analysis.
            </p>
        </div>
    </div>
    """  , unsafe_allow_html=True)
    
    st.markdown("""
    Run a custom prompt against document chunks and export the results to Excel.
    """)
    
    # Check if vector database exists
    if not VECTORDB_PATH or not os.path.exists(VECTORDB_PATH):
        st.error("No vector database found. Please ingest documents first by going to Manage.")
        return
    
    # Initialize the vector store using LLM.load_vectorstore()
    try:
        # Load the vector store
        vectorstore = llm.load_vectorstore()
        store_type = llm.get_store_type()
        
        # Check if store exists and has documents
        if not vectorstore.exists():
            st.error("No documents have been indexed. Please ingest documents first by going to Manage.")
            return
            
        # Determine if keyword search is available
        # Only allow keyword search if vectorstore is NOT a DenseStore (i.e., it's a SparseStore or DualStore)
        has_keyword_search = not isinstance(vectorstore, DenseStore)
            
    except Exception as e:
        st.error(f"Error loading search index: {str(e)}")
        return
    
    # Initialize required session state with default values if not already set
    if 'docs_for_analysis' not in st.session_state:
        st.session_state.docs_for_analysis = None
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = None
    if 'analysis_in_progress' not in st.session_state:
        st.session_state.analysis_in_progress = False
    if 'current_index' not in st.session_state:
        st.session_state.current_index = 0
    if 'total_chunks' not in st.session_state:
        st.session_state.total_chunks = 0
    if 'analysis_complete' not in st.session_state:
        st.session_state.analysis_complete = False
    
    # Get folders from the resolved DOCUMENTS_PATH
    folder_options = ["All folders"]
    
    # Use DOCUMENTS_PATH as our source for folder listing
    if DOCUMENTS_PATH and os.path.exists(DOCUMENTS_PATH):
        # Get all top-level folders
        subfolders = [d for d in os.listdir(DOCUMENTS_PATH) 
                     if os.path.isdir(os.path.join(DOCUMENTS_PATH, d))]
        folder_options.extend(subfolders)
    
    # Display completed analysis results if available
    if st.session_state.analysis_complete and st.session_state.analysis_results:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.success(f"Analysis complete! Processed {len(st.session_state.analysis_results)} document chunks.")
        
        with col2:
            # Generate Excel file and prepare download
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            results_df = pd.DataFrame(st.session_state.analysis_results)
            excel_output = prepare_excel_file(results_df)
            
            # Create download button
            st.download_button(
                label="Download Results as Excel",
                data=excel_output,
                file_name=f"document_analysis_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_button"
            )

        # Display results preview
        st.subheader("Results Preview")
        st.dataframe(results_df, use_container_width=True)
        
        # Add option to start a new analysis
        if st.button("Start New Analysis", key="new_analysis_button"):
            st.session_state.reset_triggered = True
            st.rerun()
            
        # Early return to skip the rest of the UI
        return
        
    # Document Selection section
    st.subheader("Document Selection")
    st.markdown("Select documents by folder or search query:")
    
    # Create selection interface
    col1, col2 = st.columns([3, 1])
    
    with col1:
        query = st.text_input(
            "Enter search terms:", 
            placeholder="Enter keywords to search for",
            key="query_input"
        )
    
    with col2:
        # Only show "Keyword" option if the vectorstore supports it
        search_options = ["Semantic"]
        if has_keyword_search:
            search_options.insert(0, "Keyword")
            
        # Default to first option in the list
        default_index = 0
            
        search_type = st.selectbox(
            "Search type:", 
            search_options, 
            index=default_index,
            key="search_type"
        )
    
    # Folder selection dropdown below the search box
    selected_folder = st.selectbox(
        "Folder to analyze:", 
        folder_options,
        index=0,
        key="folder_selector",
        help="Select a specific folder to analyze, or 'All folders' to search across all documents"
    )
    
    # Filters section (collapsible)
    with st.expander("Advanced Filters and Search Settings"):
        filter_options = {} # not currently implemented
        
        # Custom where clause
        where_document = st.text_input(
            "Custom query filter:", 
            placeholder='e.g., "ChatGPT" AND extension:pdf',
            help="Use AND, OR, NOT operators for complex queries",
            key="where_document"
        )
        
        # Search settings
        st.subheader("Analysis Settings")
        st.info('💡 OnPrem splits documents into passages (or chunks) during ingestion.')
        results_limit = st.slider(
            "Maximum number of passages to analyze:", 
            min_value=100, 
            max_value=5000, 
            value=1000, 
            step=100,
            key="results_limit"
        )
    
    # Prompt template section
    st.subheader("Analysis Prompt")
    st.markdown("Enter the prompt to apply to each document chunk. Use `{text}` as a placeholder for the document content.")
    
    # Default prompt template with {text} placeholder
    default_prompt = """Provide a single short keyword or keyphrase that captures the topic of the following text (only output the keyphrase without quotes and nothing else): {text}"""
    
    prompt_template = st.text_area(
        "Prompt template:",
        value=default_prompt,
        height=100,
        key="prompt_template",
        help="Use {text} as a placeholder for the document chunk content."
    )
    
    # Validate that the prompt contains the {text} placeholder
    if "{text}" not in prompt_template or not prompt_template:
        st.warning("Your prompt must be non-empty and contain the placeholder {text} to be replaced with document content.")
    
    # Action buttons
    button_cols = st.columns([1, 1, 4])  # First two columns for buttons, third for spacing
    with button_cols[0]:
        analyze_button = st.button("Analyze", type="primary", use_container_width=True)
    with button_cols[1]:
        reset_button = st.button("Reset", type="secondary", use_container_width=True)
    
    # Reset state if reset button is clicked - simplest approach is to set a flag and rerun
    if reset_button:
        st.session_state.reset_triggered = True
        st.rerun()
        
    # Process analysis when analyze button is clicked
    if analyze_button:
        # Allow empty search if a folder is selected - this will show all files in that folder
        if not query and selected_folder == "All folders":
            st.error("Please enter a search term or select a specific folder.")
            st.stop()
            
        if "{text}" not in prompt_template:
            st.error("Your prompt must contain the placeholder {text} to be replaced with document content.")
            st.stop()
            
        # Set flags to indicate analysis is starting
        st.session_state.analysis_in_progress = True
        st.session_state.current_index = 0
        st.session_state.analysis_results = []
        st.session_state.analysis_complete = False
        
        # Perform document retrieval
        with st.spinner("Retrieving document chunks..."):
            try:
                # Apply folder filtering if a specific folder is selected
                folder_filter = None
                folder_name = None
                where_clause = where_document
                
                if selected_folder != "All folders":
                    # Get folder information we want to filter by
                    folder_name = selected_folder
                    folder_path = os.path.join(DOCUMENTS_PATH, folder_name)
                    
                    # Normalize the folder path for consistent handling
                    norm_folder_path = os.path.normpath(folder_path).replace('\\', '/')
                    
                    # For more precision, we need to ensure the folder appears as a proper path component
                    folder_filter = f'source:{norm_folder_path}*'
                    
                    # Combine with existing where_document if any
                    if where_clause:
                        where_clause = f"({where_clause}) AND {folder_filter}"
                    else:
                        where_clause = folder_filter
                
                # Handle empty query case (browsing a folder)
                query_text = query
                
                # If no search query but a folder is selected, use a wildcard query
                # that will match everything in that folder
                if not query_text and selected_folder != "All folders":
                    query_text = "*"  # Wildcard to match everything
                
                if search_type == "Keyword":
                    # Use query method of vectorstore for keyword search
                    results = vectorstore.query(
                        q=query_text,
                        limit=results_limit,
                        filters=filter_options,
                        where_document=where_clause if where_clause else None,
                    )
                    hits = results.get('hits', [])
                    total_hits = results.get('total_hits', 0)
                else:  # Semantic search
                    # Use semantic_search method of vectorstore
                    if store_type != 'sparse':
                        # Transform Whoosh query filter to Chroma syntax
                        chroma_filters = lucene_to_chroma(where_clause)
                        where_document = chroma_filters['where_document']
                        filter_options = chroma_filters['filter']
                    else:
                        where_document = where_clause
                    
                    # For empty query + folder selection with semantic search,
                    # we need a special approach since semantic search needs actual text
                    if not query and selected_folder != "All folders":
                        # For semantic search with empty query, use a neutral query that will
                        # retrieve documents based primarily on the filter
                        query_text = "document"
                        
                    # Execute the semantic search
                    hits = vectorstore.semantic_search(
                        query=query_text,
                        k=results_limit,
                        filters=filter_options if filter_options else None,
                        where_document=where_document if where_document else None
                    )
                    
                    # Apply additional folder filtering for semantic search if needed
                    if selected_folder != "All folders" and query:
                        folder_path = os.path.join(DOCUMENTS_PATH, selected_folder)
                        norm_folder_path = os.path.normpath(folder_path).replace('\\', '/')
                        # Filter to only keep hits that are from the selected folder
                        hits = [hit for hit in hits if 'source' in hit.metadata and 
                                hit.metadata['source'].startswith(norm_folder_path)]
                    
                    total_hits = len(hits)
                
                # Store the documents for processing
                st.session_state.docs_for_analysis = hits
                st.session_state.total_chunks = len(hits)
                
                if total_hits == 0:
                    st.warning("No document chunks found matching your criteria.")
                    st.session_state.analysis_in_progress = False
                    return
                
                st.info(f"Found {total_hits} document chunks to analyze. Processing will begin shortly...")
                st.rerun()
                
            except Exception as e:
                st.error(f"Error retrieving documents: {str(e)}")
                st.session_state.analysis_in_progress = False
                return
    
    # Process documents if analysis is in progress
    if st.session_state.analysis_in_progress and hasattr(st.session_state, 'docs_for_analysis') and st.session_state.docs_for_analysis:
        # Create a progress bar
        progress_bar = st.progress(0)
        
        # Create a status message
        status_message = st.empty()
        
        # Load the LLM for processing
        llm_instance = load_llm()
        
        try:
            # Get the total number of documents to process
            total_docs = len(st.session_state.docs_for_analysis)
            current_index = st.session_state.current_index
            
            # Process documents in batches to allow UI updates
            batch_size = 5  # Process 5 documents at a time
            end_index = min(current_index + batch_size, total_docs)
            
            # Update status message
            status_message.info(f"Processing documents {current_index+1} to {end_index} of {total_docs}...")
            
            # Process the current batch
            for i in range(current_index, end_index):
                doc = st.session_state.docs_for_analysis[i]
                source = doc.metadata.get("source", "Unknown source")
                text = doc.page_content
                
                # Format the prompt by replacing {text} with the document content
                formatted_prompt = prompt_template.replace("{text}", text)
                
                # Process the document with the LLM
                result = llm_instance.prompt(formatted_prompt)
                
                # Store the result
                if not hasattr(st.session_state, 'analysis_results') or st.session_state.analysis_results is None:
                    st.session_state.analysis_results = []
                    
                st.session_state.analysis_results.append({
                    "result": result,
                    "source": source,
                    "text": text
                })
                
                # Update progress
                progress = (i + 1) / total_docs
                progress_bar.progress(progress)
                
            # Update the current index
            st.session_state.current_index = end_index
            
            # Check if we're done
            if end_index >= total_docs:
                st.session_state.analysis_in_progress = False
                st.session_state.analysis_complete = True
                progress_bar.progress(1.0)
                status_message.success(f"Analysis complete! Processed {total_docs} document chunks.")
                st.rerun()  # Force a rerun to show results properly
            else:
                # Still processing, rerun to continue
                st.rerun()
                
        except Exception as e:
            # Handle any errors
            st.error(f"Error processing documents: {str(e)}")
            st.session_state.analysis_in_progress = False
            
            # If we've processed some documents, still offer to download partial results
            if hasattr(st.session_state, 'analysis_results') and st.session_state.analysis_results:
                # Convert results to DataFrame
                results_df = pd.DataFrame(st.session_state.analysis_results)
                excel_output = prepare_excel_file(results_df)
                
                # Generate a timestamp for the filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                st.warning("Analysis was interrupted, but you can download the partial results:")
                
                # Create download button
                st.download_button(
                    label="Download Partial Results as Excel",
                    data=excel_output,
                    file_name=f"document_analysis_partial_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_partial_button"
                )


if __name__ == "__main__":
    # Set page to wide mode when run directly
    st.set_page_config(
        page_title="Document Analysis", 
        page_icon="📊", 
        layout="wide",
        initial_sidebar_state="expanded"
    )
    main()
