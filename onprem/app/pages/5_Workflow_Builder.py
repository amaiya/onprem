import os
import sys
import json
import yaml
import tempfile
import streamlit as st
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
from io import BytesIO
from datetime import datetime

# Add parent directory to path to allow importing when run directly
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

# Import from parent modules
from OnPrem import read_config
from utils import load_llm

# Import workflow engine
try:
    from onprem.workflow import WorkflowEngine, NODE_REGISTRY
    from onprem.workflow.exceptions import WorkflowValidationError, NodeExecutionError
    WORKFLOW_AVAILABLE = True
except ImportError:
    WORKFLOW_AVAILABLE = False

def get_valid_next_nodes(current_workflow: List[Dict] = None):
    """Get valid next nodes based on current workflow state using workflow engine logic."""
    from onprem.workflow.base import NODE_TYPES
    from onprem.workflow.registry import NODE_REGISTRY
    
    # If no nodes yet, start with Query nodes (since we work with existing vector stores)
    if not current_workflow:
        return ["Query"]
    else:
        # Get the last node in the workflow
        last_node = current_workflow[-1]
        last_node_name = last_node['type']
        
        # Get the actual node class and create a temporary instance to check its NODE_TYPE
        if last_node_name in NODE_REGISTRY:
            last_node_class = NODE_REGISTRY[last_node_name]
            last_node_type = last_node_class.NODE_TYPE
            
            # Get valid connection targets from NODE_TYPES
            node_type_info = NODE_TYPES.get(last_node_type)
            return node_type_info.can_connect_to if node_type_info else []
        else:
            return []

def get_all_node_definitions():
    """Get complete node definitions without filtering - used for templates and node data."""
    # Detect current web app store configuration
    cfg, _ = read_config()
    store_type = cfg.get("llm", {}).get("store_type", "dense")
    
    # Define all possible query nodes
    all_query_nodes = {
        'QueryDualStore': {
            'description': 'Query dual vector store (semantic + keyword)',
            'inputs': {},
            'outputs': {'documents': 'List[Document]'},
            'config_fields': {
                'query': {'type': 'text', 'required': True, 'help': 'Search query. To search documents in specific folder, choose "sparse" as search type (if availabe) with query: source:"/my/folder*"'},
                'limit': {'type': 'number', 'default': 10, 'help': 'Maximum results'},
                'search_type': {'type': 'select', 'options': ['sparse', 'semantic', 'hybrid'], 'default': 'sparse'},
                'dense_weight': {'type': 'number', 'default': 0.6, 'help': 'Weight for semantic search (0.0-1.0)', 'conditional': 'search_type', 'show_when': 'hybrid'},
                'sparse_weight': {'type': 'number', 'default': 0.4, 'help': 'Weight for sparse search (0.0-1.0)', 'conditional': 'search_type', 'show_when': 'hybrid'}
            }
        },
        'QueryWhooshStore': {
            'description': 'Query Whoosh vector store (keyword search)',
            'inputs': {},
            'outputs': {'documents': 'List[Document]'},
            'config_fields': {
                'query': {'type': 'text', 'required': True, 'help': 'Search query'},
                'limit': {'type': 'number', 'default': 10, 'help': 'Maximum results'},
                'search_type': {'type': 'select', 'options': ['sparse', 'semantic'], 'default': 'semantic'}
            }
        },
        'QueryChromaStore': {
            'description': 'Query Chroma vector store (semantic search)',
            'inputs': {},
            'outputs': {'documents': 'List[Document]'},
            'config_fields': {
                'query': {'type': 'text', 'required': True, 'help': 'Search query'},
                'limit': {'type': 'number', 'default': 10, 'help': 'Maximum results'}
            }
        },
        'QueryElasticsearchStore': {
            'description': 'Query Elasticsearch vector store (hybrid search)',
            'inputs': {},
            'outputs': {'documents': 'List[Document]'},
            'config_fields': {
                'query': {'type': 'text', 'required': True, 'help': 'Search query'},
                'limit': {'type': 'number', 'default': 10, 'help': 'Maximum results'},
                'search_type': {'type': 'select', 'options': ['sparse', 'semantic', 'hybrid'], 'default': 'hybrid'}
            }
        }
    }
    
    # Filter query nodes based on web app store configuration
    query_nodes = {}
    if store_type == "dual":
        query_nodes['QueryDualStore'] = all_query_nodes['QueryDualStore']
    elif store_type == "sparse":
        query_nodes['QueryWhooshStore'] = all_query_nodes['QueryWhooshStore']
    elif store_type in ["dense", "chroma"]:
        query_nodes['QueryChromaStore'] = all_query_nodes['QueryChromaStore']
    elif store_type == "elasticsearch":
        query_nodes['QueryElasticsearchStore'] = all_query_nodes['QueryElasticsearchStore']
    else:
        # Default fallback to dual if unknown store type
        query_nodes['QueryDualStore'] = all_query_nodes['QueryDualStore']
    
    return {
        # Query nodes (starting points) - filtered by web app configuration
        'Query': query_nodes,
        # Document transformers
        'Document Transformers': {
            'AddMetadata': {
                'description': 'Add custom metadata fields to documents',
                'inputs': {'documents': 'List[Document]'},
                'outputs': {'documents': 'List[Document]'},
                'config_fields': {
                    'metadata': {'type': 'json', 'required': True, 'help': 'JSON object with metadata to add (e.g., {"category": "research", "priority": "high"})'}
                }
            },
            'ContentPrefix': {
                'description': 'Add text to the beginning of document content',
                'inputs': {'documents': 'List[Document]'},
                'outputs': {'documents': 'List[Document]'},
                'config_fields': {
                    'prefix': {'type': 'textarea', 'required': True, 'help': 'Text to prepend to each document'},
                    'separator': {'type': 'text', 'default': '\\n\\n', 'help': 'Separator between prefix and content (use \\n for newlines, \\t for tabs)'}
                }
            },
            'ContentSuffix': {
                'description': 'Add text to the end of document content',
                'inputs': {'documents': 'List[Document]'},
                'outputs': {'documents': 'List[Document]'},
                'config_fields': {
                    'suffix': {'type': 'textarea', 'required': True, 'help': 'Text to append to each document'},
                    'separator': {'type': 'text', 'default': '\\n\\n', 'help': 'Separator between content and suffix (use \\n for newlines, \\t for tabs)'}
                }
            },
            'DocumentFilter': {
                'description': 'Filter documents by content or metadata criteria',
                'inputs': {'documents': 'List[Document]'},
                'outputs': {'documents': 'List[Document]'},
                'config_fields': {
                    'content_contains': {'type': 'text', 'help': 'Comma-separated terms that content must contain (any)'},
                    'content_excludes': {'type': 'text', 'help': 'Comma-separated terms that content must not contain (any)'},
                    'min_length': {'type': 'number', 'help': 'Minimum content length in characters'},
                    'max_length': {'type': 'number', 'help': 'Maximum content length in characters'},
                    'metadata_filters': {'type': 'json', 'help': 'JSON object with metadata filters (e.g., {"category": "research"})'}
                }
            },
            'PythonDocumentTransformer': {
                'description': 'Transform documents using custom Python code',
                'inputs': {'documents': 'List[Document]'},
                'outputs': {'documents': 'List[Document]'},
                'config_fields': {
                    'code': {
                        'type': 'textarea', 
                        'required': True, 
                        'help': 'Python code to transform documents. Available variables: doc, content, metadata, document_id, source',
                        'default': '''# PythonDocumentTransformer: Modify document content and metadata
# Available variables: doc, content, metadata, document_id, source
# Pre-available modules: re, json, math, datetime (no import needed)
# Output: Modified doc object (transforms documents in place)

#----------------
# EXAMPLE
#----------------
# Clean up content using pre-available re module
content = re.sub(r'\\s+', ' ', content).strip()

# Add metadata
metadata['word_count'] = len(content.split())
metadata['doc_type'] = 'email' if '@' in content else 'general'
metadata['processed_at'] = datetime.datetime.now().isoformat()

# Update document content
doc.page_content = content'''
                    }
                }
            }
        },
        # Document processors
        'Processors': {
            'PromptProcessor': {
                'description': 'Apply LLM prompt to each document in search results',
                'inputs': {'documents': 'List[Document]'},
                'outputs': {'results': 'List[Dict]'},
                'config_fields': {
                    'prompt': {
                        'type': 'textarea', 
                        'required': True, 
                        'help': 'LLM prompt template. Use {content} for document text.',
                        'default': '''Provide a single short keyword or keyphrase that captures the topic of the following text: {content} '''
                    },
                    'batch_size': {'type': 'number', 'default': 5, 'help': 'Documents per batch'}
                }
            },
            'SummaryProcessor': {
                'description': 'Generate document summaries for each document in search results',
                'inputs': {'documents': 'List[Document]'},
                'outputs': {'results': 'List[Dict]'},
                'config_fields': {
                    'batch_size': {'type': 'number', 'default': 5, 'help': 'Documents per batch'}
                }
            },
            'ResponseCleaner': {
                'description': 'Clean and refine LLM responses (e.g., place after PromptProcessor)',
                'inputs': {'results': 'List[Dict]'},
                'outputs': {'results': 'List[Dict]'},
                'config_fields': {
                    'source_field': {
                        'type': 'text',
                        'default': 'response',
                        'help': 'Field name to clean (e.g., "response", "page_content", "output")'
                    },
                    'cleanup_prompt': {
                        'type': 'textarea', 
                        'required': True, 
                        'help': 'Prompt to clean/refine content. Use {NAME_OF_SOURCE_FIELD} as the variable in the prompt.',
                        'default': '''Clean up and standardize this content by:
1. Removing any unnecessary formatting or markdown
2. Correcting grammar and spelling errors  
3. Making the language more concise and professional
4. Ensuring consistent formatting

Original content: {response}

Provide the cleaned version:'''
                    }
                }
            },
            'DocumentToResults': {
                'description': 'Directly convert documents to results for aggregation or export.',
                'inputs': {'documents': 'List[Document]'},
                'outputs': {'results': 'List[Dict]'},
                'config_fields': {
                    'include_content': {'type': 'checkbox', 'default': True, 'help': 'Include document content'},
                    'content_field': {'type': 'text', 'default': 'page_content', 'help': 'Name for content field'},
                    'metadata_prefix': {'type': 'text', 'default': 'meta_', 'help': 'Prefix for metadata fields'}
                }
            },
            'PythonDocumentProcessor': {
                'description': 'Process documents using custom Python code (returns results format)',
                'inputs': {'documents': 'List[Document]'},
                'outputs': {'results': 'List[Dict]'},
                'config_fields': {
                    'code': {
                        'type': 'textarea', 
                        'required': True, 
                        'help': 'Python code to process documents. Available variables: doc, content, metadata, document_id, source, result. Populate the result dict.',
                        'default': '''# PythonDocumentProcessor: Convert documents to structured results
# Available variables: doc, content, metadata, document_id, source, result (empty dict to populate)
# Pre-available modules: re, json, math, datetime (no import needed)
# Output: Populate the result dict with extracted information

#----------------
# EXAMPLE
#----------------

# Extract information using pre-available modules
emails = re.findall(r'\\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\\.[A-Z|a-z]{2,}\\b', content)
phone_numbers = re.findall(r'\\b\\d{3}-\\d{3}-\\d{4}\\b', content)

# Populate the result dictionary for this document
result['source'] = source
result['text_length'] = len(content)
result['word_count'] = len(content.split())
result['has_email'] = len(emails) > 0
result['has_phone'] = len(phone_numbers) > 0
result['timestamp'] = datetime.datetime.now().isoformat()
result['preview'] = content[:100] + '...' if len(content) > 100 else content'''
                    }
                }
            },
            'PythonResultProcessor': {
                'description': 'Process results using custom Python code',
                'inputs': {'results': 'List[Dict]'},
                'outputs': {'results': 'List[Dict]'},
                'config_fields': {
                    'code': {
                        'type': 'textarea', 
                        'required': True, 
                        'help': 'Python code to process results. Available variables: result, result_id. Modify result dict in place or set new_result.',
                        'default': '''# PythonResultProcessor: Enhance and post-process results
# Available variables: result (dict), result_id
# Pre-available modules: re, json, math, datetime (no import needed)
# Output: Modify result dict in place or set new_result

#----------------
# EXAMPLE
#----------------

# Add timestamp and metadata  
result['processed_at'] = datetime.datetime.now().isoformat()

# Add confidence score based on text length
if 'text_length' in result:
    result['confidence'] = 'high' if result['text_length'] > 1000 else 'low'

# Extract filename from source using string operations
if 'source' in result:
    source_parts = result['source'].split('/')
    result['filename'] = source_parts[-1] if source_parts else 'unknown'

# Add category based on content analysis
if 'has_email' in result and result['has_email']:
    result['category'] = 'communication'
else:
    result['category'] = 'document' '''
                    }
                }
            }
        },
        # Aggregators
        'Aggregators': {
            'AggregatorNode': {
                'description': 'Aggregate multiple results with LLM',
                'inputs': {'results': 'List[Dict]'},
                'outputs': {'result': 'Dict'},
                'config_fields': {
                    'prompt': {
                        'type': 'textarea', 
                        'required': True, 
                        'help': 'Prompt to combine all previous results into a single summary. Use {responses} for all the individual results.',
                        'default': '''You are analyzing multiple document analysis results. Please create a comprehensive summary that combines all the insights.

Individual Results:
{responses}

Please provide a consolidated analysis including:
1. **Overall Themes**: What are the main topics across all documents?
2. **Key Findings**: What are the most important insights or patterns?
3. **Common Elements**: What appears frequently across the results?
4. **Summary Statistics**: How many documents, main categories, etc.
5. **Conclusions**: What can we conclude from this collection of documents?

Format as a well-structured summary report.'''
                    }
                }
            },
            'PythonAggregatorNode': {
                'description': 'Aggregate results with Python code',
                'inputs': {'results': 'List[Dict]'},
                'outputs': {'result': 'Dict'},
                'config_fields': {
                    'code': {
                        'type': 'textarea', 
                        'required': True, 
                        'help': 'Python aggregation code. Available variables: results (List[Dict]). Set result dict.',
                        'default': '''# PythonAggregatorNode: Combine multiple results into single summary
# Available variables: results (List[Dict])
# Pre-available modules: re, json, math, datetime (no import needed)
# Output: Set aggregated_result dict

#----------------
# EXAMPLE
#----------------

# Create aggregated summary
result = {
    'timestamp': datetime.datetime.now().isoformat(),
    'total_results': len(results),
    'total_length': sum(r.get('text_length', 0) for r in results),
    'sources': [r.get('source', 'Unknown') for r in results],
    'high_confidence_count': sum(1 for r in results if r.get('confidence') == 'high'),
    'average_word_count': sum(r.get('word_count', 0) for r in results) / len(results) if results else 0
}'''
                    }
                }
            }
        },
        # Exporters
        'Exporters': {
            'CSVExporter': {
                'description': 'Export results to CSV',
                'inputs': {'results': 'List[Dict]'},
                'outputs': {'status': 'str'},
                'config_fields': {
                    'output_path': {'type': 'text', 'default': 'results.csv', 'help': 'Output file path'}
                }
            },
            'ExcelExporter': {
                'description': 'Export results to Excel spreadsheet',
                'inputs': {'results': 'List[Dict]'},
                'outputs': {'status': 'str'},
                'config_fields': {
                    'output_path': {'type': 'text', 'default': 'results.xlsx', 'help': 'Output file path'},
                    'sheet_name': {'type': 'text', 'default': 'Results', 'help': 'Name of the Excel sheet'}
                }
            },
            'JSONExporter': {
                'description': 'Export results to JSON',
                'inputs': {'results': 'List[Dict]', 'result': 'Dict'},
                'outputs': {'status': 'str'},
                'config_fields': {
                    'output_path': {'type': 'text', 'default': 'results.json', 'help': 'Output file path'},
                    'pretty_print': {'type': 'checkbox', 'default': True, 'help': 'Format JSON nicely'}
                }
            },
            'JSONResponseExporter': {
                'description': 'Extract and export JSON from responses',
                'inputs': {'results': 'List[Dict]', 'result': 'Dict'},
                'outputs': {'status': 'str'},
                'config_fields': {
                    'output_path': {'type': 'text', 'default': 'extracted_responses.json', 'help': 'Output file path'},
                    'response_field': {'type': 'text', 'default': 'response', 'help': 'Field containing JSON response'}
                }
            }
        }
    }

def get_available_nodes(current_workflow: List[Dict] = None):
    """Get nodes that work with existing vector stores, filtered by workflow state."""
    
    # Get valid next node types based on workflow state
    valid_node_types = get_valid_next_nodes(current_workflow)
    
    # Get all node definitions
    all_nodes = get_all_node_definitions()
    
    # Filter nodes based on valid next node types using workflow engine logic
    from onprem.workflow.registry import NODE_REGISTRY
    
    filtered_categories = {}
    
    # Go through each category
    for category_name, category_nodes in all_nodes.items():
        # Skip Exporters category - results are automatically available for download
        if category_name == 'Exporters':
            continue
            
        filtered_category_nodes = {}
        
        # Filter nodes within this category
        for node_name, node_config in category_nodes.items():
            # Get the node class and check its NODE_TYPE
            if node_name in NODE_REGISTRY:
                node_class = NODE_REGISTRY[node_name]
                node_type = node_class.NODE_TYPE
                
                # Only include if this node type is valid for the current workflow state
                if node_type in valid_node_types:
                    # Additional port-level validation for processors
                    if current_workflow and node_type == "Processor":
                        last_node = current_workflow[-1]
                        last_node_name = last_node['type']
                        
                        # Check port compatibility for processor connections
                        if last_node_name in NODE_REGISTRY:
                            last_node_class = NODE_REGISTRY[last_node_name]
                            
                            # Create temporary instances to get type information
                            try:
                                last_node_instance = last_node_class("temp_last", {})
                                current_node_instance = node_class("temp_current", {})
                                
                                # Get output and input types using the actual methods
                                last_outputs = last_node_instance.get_output_types()
                                current_inputs = current_node_instance.get_input_types()
                                
                                # Check if any output from last node can connect to any input of current node
                                compatible = False
                                for output_port, output_type in last_outputs.items():
                                    for input_port, input_type in current_inputs.items():
                                        if output_type == input_type:
                                            compatible = True
                                            break
                                    if compatible:
                                        break
                                
                                # Skip this node if no compatible connection exists
                                if not compatible:
                                    continue
                            except Exception:
                                # If we can't create instances, skip validation and allow connection
                                pass
                    
                    filtered_category_nodes[node_name] = node_config
        
        # Only include category if it has any valid nodes
        if filtered_category_nodes:
            filtered_categories[category_name] = filtered_category_nodes
    
    return filtered_categories

def render_node_config(node_type: str, node_data: Dict, node_id: str, existing_config: Dict = None) -> Dict:
    """Render configuration UI for a node and return config"""
    config = {}
    existing_config = existing_config or {}
    
    st.subheader(f"{node_type} Configuration")
    st.write(f"**{node_id}**: {node_data['description']}")
    
    # LLM nodes automatically use the web app's configured LLM
    if node_type in ['PromptProcessor', 'SummaryProcessor', 'ResponseCleaner', 'AggregatorNode']:
        st.info("💡 This node will automatically use the same LLM configured for the web app.")
    
    # Automatically set vector store path for query nodes from web app configuration
    if node_type.startswith('Query'):
        cfg, _ = read_config()
        vectordb_path = cfg.get('vectordb_path', '~/onprem_data/webapp/vectordb')
        config['persist_location'] = vectordb_path
        st.info(f"🔗 Vector store path automatically set to: `{vectordb_path}`")
    
    # Render node-specific config fields
    for field_name, field_config in node_data['config_fields'].items():
        key = f"{node_id}_{field_name}"
        
        # Check if this field should be conditionally displayed
        should_show = True
        if 'conditional' in field_config and 'show_when' in field_config:
            conditional_field = field_config['conditional']
            show_when_value = field_config['show_when']
            
            # Get the value of the conditional field from the current config
            conditional_key = f"{node_id}_{conditional_field}"
            if conditional_key in st.session_state:
                current_value = st.session_state[conditional_key]
                should_show = (current_value == show_when_value)
            else:
                # If conditional field hasn't been set yet, check default value
                conditional_field_config = node_data['config_fields'].get(conditional_field, {})
                default_value = conditional_field_config.get('default', '')
                should_show = (default_value == show_when_value)
        
        if not should_show:
            continue
            
        if field_config['type'] == 'text':
            # Use existing config value if available, otherwise use field default
            default_value = existing_config.get(field_name, field_config.get('default', ''))
            value = st.text_input(
                field_name.replace('_', ' ').title(),
                value=default_value,
                key=key,
                help=field_config.get('help', '')
            )
            if value or field_config.get('required'):
                # Process escape sequences for separator fields
                if field_name == 'separator' and value:
                    # Decode common escape sequences
                    processed_value = value.replace('\\n', '\n').replace('\\t', '\t').replace('\\r', '\r')
                    config[field_name] = processed_value
                else:
                    config[field_name] = value
                
        elif field_config['type'] == 'textarea':
            # Use existing config value if available, otherwise use field default
            default_value = existing_config.get(field_name, field_config.get('default', ''))
            value = st.text_area(
                field_name.replace('_', ' ').title(),
                value=default_value,
                height=150,
                key=key,
                help=field_config.get('help', '')
            )
            if value or field_config.get('required'):
                config[field_name] = value
                
        elif field_config['type'] == 'number':
            # Special handling for weights (0.0-1.0 range)
            if 'weight' in field_name:
                # Use existing config value if available, otherwise use field default
                default_value = existing_config.get(field_name, field_config.get('default', 0.5))
                value = st.slider(
                    field_name.replace('_', ' ').title(),
                    min_value=0.0,
                    max_value=1.0,
                    value=float(default_value),
                    step=0.1,
                    key=key,
                    help=field_config.get('help', '')
                )
            else:
                # Use existing config value if available, otherwise use field default
                default_value = existing_config.get(field_name, field_config.get('default', 1))
                value = st.number_input(
                    field_name.replace('_', ' ').title(),
                    value=int(default_value),
                    min_value=1,
                    key=key,
                    help=field_config.get('help', '')
                )
            config[field_name] = value
            
        elif field_config['type'] == 'checkbox':
            # Use existing config value if available, otherwise use field default
            default_value = existing_config.get(field_name, field_config.get('default', False))
            value = st.checkbox(
                field_name.replace('_', ' ').title(),
                value=bool(default_value),
                key=key,
                help=field_config.get('help', '')
            )
            config[field_name] = value
            
        elif field_config['type'] == 'select':
            # Use existing config value if available, otherwise use field default
            default_value = existing_config.get(field_name, field_config.get('default', field_config['options'][0]))
            try:
                default_index = field_config['options'].index(default_value)
            except ValueError:
                # If existing value is not in options, fall back to field default
                default_value = field_config.get('default', field_config['options'][0])
                default_index = field_config['options'].index(default_value)
            
            value = st.selectbox(
                field_name.replace('_', ' ').title(),
                options=field_config['options'],
                index=default_index,
                key=key,
                help=field_config.get('help', '')
            )
            config[field_name] = value
            
        elif field_config['type'] == 'json':
            import json
            placeholder_text = field_config.get('help', 'Enter JSON object')
            
            # Use existing config value if available, otherwise use field default
            if field_name in existing_config:
                # If we have existing config, use it (may be dict or string)
                existing_value = existing_config[field_name]
                if isinstance(existing_value, dict):
                    default_text = json.dumps(existing_value, indent=2)
                else:
                    default_text = str(existing_value)
            else:
                # Fall back to field default
                field_default = field_config.get('default', '{}')
                default_text = field_default if isinstance(field_default, str) else json.dumps(field_default, indent=2)
            
            value = st.text_area(
                field_name.replace('_', ' ').title(),
                value=default_text,
                height=100,
                key=key,
                help=placeholder_text,
                placeholder='{"key": "value", "number": 123}'
            )
            
            # Validate and parse JSON
            if value.strip():
                try:
                    parsed_value = json.loads(value)
                    config[field_name] = parsed_value
                    # Clear any previous error state
                    if f"{key}_error" in st.session_state:
                        del st.session_state[f"{key}_error"]
                except json.JSONDecodeError as e:
                    st.error(f"Invalid JSON in {field_name}: {str(e)}")
                    st.session_state[f"{key}_error"] = True
                    # Don't include invalid JSON in config
            elif field_config.get('required'):
                st.error(f"{field_name} is required")
    
    return config

def load_workflow_from_yaml(yaml_content: str) -> List[Dict]:
    """Parse YAML workflow and convert to workflow nodes format"""
    try:
        workflow_dict = yaml.safe_load(yaml_content)
        
        if not isinstance(workflow_dict, dict) or 'nodes' not in workflow_dict:
            raise ValueError("Invalid workflow format: missing 'nodes' section")
        
        workflow_nodes = []
        all_nodes = get_all_node_definitions()
        
        # Convert YAML nodes to workflow_nodes format
        for node_id, node_config in workflow_dict['nodes'].items():
            if 'type' not in node_config:
                raise ValueError(f"Node '{node_id}' missing 'type' field")
            
            node_type = node_config['type']
            
            # Find the node in our definitions
            found_node = None
            found_category = None
            
            for category_name, category_nodes in all_nodes.items():
                if node_type in category_nodes:
                    found_node = category_nodes[node_type]
                    found_category = category_name
                    break
            
            if not found_node:
                raise ValueError(f"Unknown node type: {node_type}")
            
            # Extract config
            config = node_config.get('config', {})
            
            # Special handling for QueryDualStore weights (convert back to individual fields)
            if node_type == 'QueryDualStore' and 'weights' in config:
                weights = config['weights']
                if isinstance(weights, list) and len(weights) >= 2:
                    config['dense_weight'] = weights[0]
                    config['sparse_weight'] = weights[1]
                config.pop('weights', None)
            
            # Special handling for DocumentFilter (convert lists back to comma-separated strings)
            if node_type == 'DocumentFilter':
                for field_name in ['content_contains', 'content_excludes']:
                    if field_name in config and isinstance(config[field_name], list):
                        config[field_name] = ', '.join(config[field_name])
            
            workflow_nodes.append({
                'id': node_id,
                'type': node_type,
                'category': found_category,
                'data': found_node,
                'config': config
            })
        
        return workflow_nodes
        
    except yaml.YAMLError as e:
        raise ValueError(f"Invalid YAML format: {str(e)}")
    except Exception as e:
        raise ValueError(f"Error parsing workflow: {str(e)}")

def generate_workflow_yaml(workflow_nodes: List[Dict]) -> str:
    """Generate YAML workflow from node configuration"""
    if not workflow_nodes:
        return "# No nodes configured"
    
    nodes = {}
    connections = []
    
    # Build nodes
    for i, node in enumerate(workflow_nodes):
        node_id = node['id']
        config = node['config'].copy()
        
        # Special handling for QueryDualStore weights
        if node['type'] == 'QueryDualStore' and 'dense_weight' in config and 'sparse_weight' in config:
            weights = [config['dense_weight'], config['sparse_weight']]
            config['weights'] = weights
            # Remove individual weight fields
            config.pop('dense_weight', None)
            config.pop('sparse_weight', None)
        
        # Special handling for DocumentFilter comma-separated lists
        if node['type'] == 'DocumentFilter':
            # Convert comma-separated strings to lists
            for field_name in ['content_contains', 'content_excludes']:
                if field_name in config and isinstance(config[field_name], str):
                    # Split by comma and strip whitespace
                    config[field_name] = [term.strip() for term in config[field_name].split(',') if term.strip()]
        
        nodes[node_id] = {
            'type': node['type'],
            'config': config
        }
        
        # Add connections (simple linear chain for now)
        if i > 0:
            prev_node = workflow_nodes[i-1]
            # Determine correct port connections based on node types
            from_port = get_output_port(prev_node['type'])
            to_port = get_input_port(node['type'])
            
            connections.append({
                'from': prev_node['id'],
                'from_port': from_port,
                'to': node_id,
                'to_port': to_port
            })
    
    workflow = {
        'nodes': nodes,
        'connections': connections
    }
    
    return yaml.dump(workflow, default_flow_style=False, sort_keys=False)

def get_output_port(node_type: str) -> str:
    """Get the primary output port for a node type"""
    if node_type.startswith('Query'):
        return 'documents'
    elif node_type in ['AddMetadata', 'ContentPrefix', 'ContentSuffix', 'DocumentFilter', 'PythonDocumentTransformer']:
        return 'documents'
    elif node_type in ['PromptProcessor', 'SummaryProcessor', 'ResponseCleaner', 'PythonDocumentProcessor', 'PythonResultProcessor', 'DocumentToResults']:
        return 'results'
    elif node_type in ['AggregatorNode', 'PythonAggregatorNode']:
        return 'result'
    else:
        return 'status'

def get_input_port(node_type: str) -> str:
    """Get the primary input port for a node type"""
    if node_type in ['PromptProcessor', 'SummaryProcessor', 'PythonDocumentProcessor', 'DocumentToResults', 'AddMetadata', 'ContentPrefix', 'ContentSuffix', 'DocumentFilter', 'PythonDocumentTransformer']:
        return 'documents'
    elif node_type in ['ResponseCleaner', 'PythonResultProcessor']:
        return 'results'
    elif node_type in ['AggregatorNode', 'PythonAggregatorNode']:
        return 'results'
    elif node_type.endswith('Exporter'):
        return 'results'
    else:
        return 'documents'

def sanitize_for_excel(text):
    """Helper function to sanitize text for Excel"""
    if isinstance(text, str):
        # Remove control characters and other characters that Excel can't handle
        chars_to_remove = '\x00-\x08\x0B-\x0C\x0E-\x1F\x7F'
        import re
        return re.sub(f'[{chars_to_remove}]', '', text)
    return text

def format_excel_writer(writer, df):
    """Apply Excel formatting for better appearance"""
    # Get the workbook and sheet
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # Create a bold format for headers
    from openpyxl.styles import Font, Alignment, PatternFill
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Apply formatting to headers
    for col_num, column_title in enumerate(df.columns):
        cell = worksheet.cell(row=1, column=col_num+1)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
    
    # Auto-adjust columns' width to fit content
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        # Add some padding (8 characters)
        adjusted_width = (length + 8)
        # Limit maximum width to avoid extremely wide columns
        column_width = min(adjusted_width, 100)
        worksheet.column_dimensions[column_cells[0].column_letter].width = column_width

def prepare_excel_file(results_df):
    """Create a formatted Excel file in memory"""
    # Apply sanitization to all text columns
    for col in results_df.columns:
        results_df[col] = results_df[col].apply(sanitize_for_excel)
    
    # Create Excel file in memory with enhanced formatting
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        results_df.to_excel(writer, index=False)
        # Apply formatting to make the Excel file more readable
        format_excel_writer(writer, results_df)
    output.seek(0)
    
    return output

def display_workflow_results(results):
    """Display workflow results following Document Analysis pattern with automatic dataframe display and Excel download"""
    if not results:
        st.info("No results to display")
        return
    
    # Get the last node's results
    if isinstance(results, dict):
        # Find the highest numbered node (e.g., node_2 is higher than node_1)
        node_keys = [k for k in results.keys() if k.startswith('node_')]
        if node_keys:
            # Sort by node number
            last_node_key = sorted(node_keys, key=lambda x: int(x.split('_')[1]))[-1]
            last_node_results = results[last_node_key]
            
            st.info(f"📊 Displaying results from final node: **{last_node_key}**")
        else:
            # If no node_ keys, just take the first key
            last_node_key = list(results.keys())[-1]
            last_node_results = results[last_node_key]
            st.info(f"📊 Displaying results from: **{last_node_key}**")
    else:
        last_node_results = results
        last_node_key = "results"
    
    # Check if the last node results contain exportable data
    exportable_data = None
    
    # Handle nested dict structure (node results contain output ports)
    if isinstance(last_node_results, dict):
        for output_key, output_value in last_node_results.items():
            if isinstance(output_value, list) and output_value:
                first_item = output_value[0]
                
                # Check if it's List[Document] - convert to List[Dict]
                if hasattr(first_item, 'page_content') and hasattr(first_item, 'metadata'):
                    exportable_data = []
                    for i, doc in enumerate(output_value):
                        doc_dict = {
                            'document_id': i + 1,
                            'content': doc.page_content,  # Full content for export
                            'source': doc.metadata.get('source', 'Unknown'),
                        }
                        # Add ALL metadata fields (like ExcelExport/CSVExport would)
                        for meta_key, meta_value in doc.metadata.items():
                            if meta_key not in doc_dict:  # Don't overwrite existing keys
                                doc_dict[meta_key] = meta_value
                        
                        exportable_data.append(doc_dict)
                    break
                
                # Check if it's already List[Dict] format
                elif isinstance(first_item, dict):
                    exportable_data = output_value
                    break
                
                # Check if it's a list of strings - convert to List[Dict]
                elif isinstance(first_item, str):
                    exportable_data = [{'result': item, 'index': i+1} for i, item in enumerate(output_value)]
                    break
            
            # Handle single dictionary result (from aggregator nodes)
            elif isinstance(output_value, dict) and output_value:
                # Convert single dict to single-row table format
                exportable_data = [output_value]
                break
    
    # Handle direct list structure
    elif isinstance(last_node_results, list) and last_node_results:
        first_item = last_node_results[0]
        
        # Check if it's List[Document] - convert to List[Dict]
        if hasattr(first_item, 'page_content') and hasattr(first_item, 'metadata'):
            exportable_data = []
            for i, doc in enumerate(last_node_results):
                doc_dict = {
                    'document_id': i + 1,
                    'content': doc.page_content,  # Full content for export
                    'source': doc.metadata.get('source', 'Unknown'),
                }
                # Add ALL metadata fields (like ExcelExport/CSVExport would)
                for meta_key, meta_value in doc.metadata.items():
                    if meta_key not in doc_dict:  # Don't overwrite existing keys
                        doc_dict[meta_key] = meta_value
                
                exportable_data.append(doc_dict)
        
        # Check if it's already List[Dict] format
        elif isinstance(first_item, dict):
            exportable_data = last_node_results
        
        # Check if it's a list of strings - convert to List[Dict]
        elif isinstance(first_item, str):
            exportable_data = [{'result': item, 'index': i+1} for i, item in enumerate(last_node_results)]
    
    # Handle single dictionary result (aggregator node output)
    elif isinstance(last_node_results, dict) and last_node_results:
        # Check if it's a nested structure with output ports
        if any(k in ['result', 'results', 'documents', 'status'] for k in last_node_results.keys()):
            # This looks like it has output ports - already handled above
            pass
        else:
            # This is a direct dict result, convert to single-row table
            exportable_data = [last_node_results]
    
    if exportable_data:
        st.subheader("Workflow Results")
        
        # Convert to DataFrame
        results_df = pd.DataFrame(exportable_data)
        
        # Display complete results (following Document Analysis pattern)
        st.subheader("Results")
        st.dataframe(results_df, use_container_width=True, height=400)
        
        # Create download button (following Document Analysis pattern)
        col1, col2 = st.columns([4, 1])
        with col2:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_output = prepare_excel_file(results_df)
            
            st.download_button(
                label="Download Results as Excel",
                data=excel_output,
                file_name=f"workflow_results_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col1:
            st.success(f"✅ Generated {len(results_df)} result records from workflow")
    
    else:
        # Display results in a more basic format if no exportable data found
        st.subheader("Workflow Results")
        st.warning("No tabular data found - displaying raw results from final node")
        
        if isinstance(last_node_results, dict):
            st.json(last_node_results)
        elif isinstance(last_node_results, list):
            for i, item in enumerate(last_node_results):
                st.write(f"{i+1}. {item}")
        else:
            st.write(str(last_node_results))

def main():
    """Visual Workflow Builder main function"""
    if not WORKFLOW_AVAILABLE:
        st.error("⚠️ Workflow engine not available. Please install the workflow dependencies.")
        return
    
    # Custom CSS for better styling
    st.markdown("""
    <style>
    .workflow-node {
        border: 2px solid #e1e5e9;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
        background-color: #f8f9fa;
    }
    .node-query { border-color: #007bff; background-color: #e7f3ff; }
    .node-processor { border-color: #28a745; background-color: #e7f8ea; }
    .node-aggregator { border-color: #ffc107; background-color: #fff8e1; }
    .node-exporter { border-color: #dc3545; background-color: #ffeaea; }
    
    .workflow-chain {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        margin: 1rem 0;
        padding: 1rem;
        background: linear-gradient(90deg, #f8f9fa 0%, #e9ecef 100%);
        border-radius: 8px;
        overflow-x: auto;
    }
    .node-box {
        min-width: 100px;
        padding: 0.5rem;
        background: white;
        border: 1px solid #dee2e6;
        border-radius: 4px;
        text-align: center;
        font-size: 0.75rem;
        white-space: nowrap;
    }
    .workflow-arrow {
        font-size: 1.2rem;
        color: #6c757d;
        margin: 0 0.25rem;
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <h1 style="
        color: #0068c9;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #0068c9;
        margin-bottom: 1.5rem;
        display: flex;
        align-items: center;
    ">
        <span style="font-size: 1.8rem; margin-right: 0.5rem;">🔧</span> 
        Visual Workflow Builder
    </h1>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    Build document intelligence workflows visually. All workflows start with a query node 
    to search your documents, then add processing and aggregation nodes to support your analysis.
    
    💡 **Quick Start:** Add a Query node → Processor node for a complete workflow. Results displayed at bottom of screen.
    """)
    
    # Initialize session state
    if 'workflow_nodes' not in st.session_state:
        st.session_state.workflow_nodes = []
    if 'node_counter' not in st.session_state:
        st.session_state.node_counter = 1
    if 'workflow_results' not in st.session_state:
        st.session_state.workflow_results = None
    
    # Get available nodes filtered by current workflow state
    available_nodes = get_available_nodes(st.session_state.workflow_nodes)
    
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.header("Add Nodes")
        
        if not available_nodes:
            st.info("✅ Workflow is complete! No additional nodes can be added.")
            st.write("Exporter nodes are terminal - they cannot connect to other nodes.")
        else:
            # Create consolidated node options with category prefixes
            # Maintain specific ordering: Query → Document Transformers → Processors → Aggregators
            category_order = ['Query', 'Document Transformers', 'Processors', 'Aggregators']
            
            consolidated_options = []
            node_mapping = {}  # Map display name to (category, node_type)
            
            # Build options in the specified order
            for category in category_order:
                if category in available_nodes:
                    for node_type, node_data in available_nodes[category].items():
                        display_name = f"{category}: {node_type}"
                        consolidated_options.append(display_name)
                        node_mapping[display_name] = (category, node_type)
            
            # Add any remaining categories not in the specified order
            for category in available_nodes.keys():
                if category not in category_order:
                    for node_type, node_data in available_nodes[category].items():
                        display_name = f"{category}: {node_type}"
                        consolidated_options.append(display_name)
                        node_mapping[display_name] = (category, node_type)
            
            # Single node selection
            selected_option = st.selectbox("Available Nodes", consolidated_options)
            
            # Get the actual category and node type
            selected_category, selected_node_type = node_mapping[selected_option]
            
            # Show node description
            node_data = available_nodes[selected_category][selected_node_type]
            st.info(f"**Description:** {node_data['description']}")
            
            # Add node button
            if st.button("➕ Add Node"):
                # Validation: Ensure first node is a Query node
                if not st.session_state.workflow_nodes and selected_category != 'Query':
                    st.error("⚠️ First node must be a Query node to search your documents!")
                else:
                    node_id = f"node_{st.session_state.node_counter}"
                    st.session_state.node_counter += 1
                    
                    # Add to workflow
                    st.session_state.workflow_nodes.append({
                        'id': node_id,
                        'type': selected_node_type,
                        'category': selected_category,
                        'data': node_data,
                        'config': {}
                    })
                    st.success(f"✅ Added {selected_node_type} node!")
                    st.rerun()
        
        # Clear workflow button
        if st.button("🗑️ Clear Workflow"):
            st.session_state.workflow_nodes = []
            st.session_state.node_counter = 1
            st.rerun()
        
        st.markdown("---")
        
        # Quick templates
        st.subheader("📋 Quick Templates")
        
        # Import workflow in a collapsible expander
        with st.expander("📥 Import YAML Workflow"):
            uploaded_file = st.file_uploader(
                "Upload YAML Workflow",
                type=['yaml', 'yml'],
                help="Upload a YAML workflow file to load it into the builder",
                key="yaml_uploader"
            )
            
            if uploaded_file is not None:
                try:
                    # Read the uploaded file
                    yaml_content = uploaded_file.read().decode('utf-8')
                    
                    # Show preview
                    st.text("📄 YAML Preview:")
                    st.code(yaml_content, language='yaml', height=200)
                    
                    # Import button
                    if st.button("🔄 Import Workflow", type="primary"):
                        try:
                            # Parse and load the workflow
                            imported_nodes = load_workflow_from_yaml(yaml_content)
                            
                            # Clear any existing node configuration from session state to avoid conflicts
                            keys_to_remove = [k for k in st.session_state.keys() if k.startswith('node_') and '_' in k[5:]]
                            for key in keys_to_remove:
                                del st.session_state[key]
                            
                            # Update session state
                            st.session_state.workflow_nodes = imported_nodes
                            st.session_state.node_counter = len(imported_nodes) + 1
                            
                            st.success(f"✅ Successfully imported workflow with {len(imported_nodes)} nodes!")
                            st.rerun()
                            
                        except ValueError as e:
                            st.error(f"❌ Import failed: {str(e)}")
                        except Exception as e:
                            st.error(f"❌ Unexpected error: {str(e)}")
                            
                except UnicodeDecodeError:
                    st.error("❌ Error reading file. Please ensure it's a valid text file.")
        
        if st.button("🚀 Document Analysis Template"):
            # Get complete node definitions for template
            all_nodes = get_all_node_definitions()
            
            # Get appropriate query node type based on store configuration
            cfg, _ = read_config()
            store_type = cfg.get("llm", {}).get("store_type", "dense")
            
            if store_type == "dual":
                query_type = 'QueryDualStore'
            elif store_type == "sparse":
                query_type = 'QueryWhooshStore'  
            elif store_type in ["dense", "chroma"]:
                query_type = 'QueryChromaStore'
            elif store_type == "elasticsearch":
                query_type = 'QueryElasticsearchStore'
            else:
                query_type = 'QueryDualStore'
            
            st.session_state.workflow_nodes = [
                {
                    'id': 'query_docs',
                    'type': query_type,
                    'category': 'Query',
                    'data': all_nodes['Query'][query_type],
                    'config': {'query': 'artificial intelligence', 'limit': 10, 'search_type': 'sparse'}
                },
                {
                    'id': 'analyze_docs',
                    'type': 'PromptProcessor',
                    'category': 'Processors',
                    'data': all_nodes['Processors']['PromptProcessor'],
                    'config': {'prompt': 'Analyze this document and extract key themes:\n\n{content}'}
                },
                # comment out as export nodes are not currently needed in Web UI
                #{
                    #'id': 'export_results',
                    #'type': 'CSVExporter',
                    #'category': 'Exporters',
                    #'data': all_nodes['Exporters']['CSVExporter'],
                    #'config': {'output_path': 'document_analysis.csv'}
                #}
            ]
            st.session_state.node_counter = 4
            st.rerun()
        
        if st.button("📊 Summary & Aggregate Template"):
            # Get complete node definitions for template
            all_nodes = get_all_node_definitions()
            
            # Get appropriate query node type based on store configuration
            cfg, _ = read_config()
            store_type = cfg.get("llm", {}).get("store_type", "dense")
            
            if store_type == "dual":
                query_type = 'QueryDualStore'
            elif store_type == "sparse":
                query_type = 'QueryWhooshStore'
            elif store_type in ["dense", "chroma"]:
                query_type = 'QueryChromaStore'
            elif store_type == "elasticsearch":
                query_type = 'QueryElasticsearchStore'
            else:
                query_type = 'QueryDualStore'
                
            st.session_state.workflow_nodes = [
                {
                    'id': 'search_docs',
                    'type': query_type,
                    'category': 'Query',
                    'data': all_nodes['Query'][query_type],
                    'config': {'query': '"generative AI"', 'limit': 5, 'search_type': 'semantic'}
                },
                {
                    'id': 'summarize',
                    'type': 'SummaryProcessor',
                    'category': 'Processors',
                    'data': all_nodes['Processors']['SummaryProcessor'],
                    'config': {}
                },
                {
                    'id': 'aggregate_summaries',
                    'type': 'AggregatorNode',
                    'category': 'Aggregators',
                    'data': all_nodes['Aggregators']['AggregatorNode'],
                    'config': {'prompt': 'Combine these summaries into a comprehensive overview:\n\n{responses}'}
                },
                # comment out as export nodes are not currently needed in Web UI
                #{
                    #'id': 'export_json',
                    #'type': 'JSONExporter',
                    #'category': 'Exporters',
                    #'data': all_nodes['Exporters']['JSONExporter'],
                    #'config': {'output_path': 'aggregated_summary.json'}
                #}
            ]
            st.session_state.node_counter = 5
            st.rerun()
    
    with col2:
        st.header("Workflow Configuration")
        
        if not st.session_state.workflow_nodes:
            st.info("👆 Add nodes from the left panel to start building your workflow")
        else:
            ## Show workflow chain
            #st.subheader("Workflow Chain")
            
            # Create visual workflow chain
            chain_html = '<div class="workflow-chain">'
            for i, node in enumerate(st.session_state.workflow_nodes):
                if i > 0:
                    chain_html += '<span class="workflow-arrow">→</span>'
                
                # Determine node color class
                if node['category'] == 'Query':
                    color_class = 'node-query'
                elif node['category'] == 'Processors':
                    color_class = 'node-processor'
                elif node['category'] == 'Aggregators':
                    color_class = 'node-aggregator'
                else:
                    color_class = 'node-exporter'
                
                chain_html += f'''
                <div class="node-box {color_class}">
                    <strong>{node['id']}</strong><br>
                    {node['type']}
                </div>
                '''
            chain_html += '</div>'
            
            st.markdown(chain_html, unsafe_allow_html=True)
            
            # Configure each node
            tabs = st.tabs([f"{node['id']}" for node in st.session_state.workflow_nodes])
            
            for i, (tab, node) in enumerate(zip(tabs, st.session_state.workflow_nodes)):
                with tab:
                    # Node configuration
                    config = render_node_config(node['type'], node['data'], node['id'], node.get('config', {}))
                    st.session_state.workflow_nodes[i]['config'] = config
                    
                    # Remove node button
                    if st.button(f"🗑️ Remove {node['id']}", key=f"remove_{node['id']}"):
                        st.session_state.workflow_nodes.pop(i)
                        st.rerun()
    
    # Generate and run workflow
    if st.session_state.workflow_nodes:
        st.header("Generated Workflow")
        
        # Generate YAML
        workflow_yaml = generate_workflow_yaml(st.session_state.workflow_nodes)
        
        # Show YAML in expandable section
        with st.expander("📄 View Generated YAML"):
            st.code(workflow_yaml, language='yaml')
        
        # Download YAML
        st.download_button(
            "💾 Download Workflow YAML",
            workflow_yaml,
            file_name="workflow.yaml",
            mime="text/yaml"
        )
        
        # Run workflow
        col1, col2 = st.columns(2)
        
        # Store results in session state to display outside columns
        if 'workflow_results' not in st.session_state:
            st.session_state.workflow_results = None
        
        with col1:
            if st.button("▶️ Run Workflow", type="primary"):
                try:
                    with st.spinner("Running workflow..."):
                        # Parse YAML and execute
                        workflow_dict = yaml.safe_load(workflow_yaml)
                        
                        engine = WorkflowEngine()
                        engine.load_workflow_from_dict(workflow_dict)
                        results = engine.execute()
                        
                        st.success("✅ Workflow completed successfully!")
                        
                        # Store results in session state
                        st.session_state.workflow_results = results
                        
                except Exception as e:
                    st.error(f"❌ Workflow execution failed: {str(e)}")
                    st.exception(e)
                    st.session_state.workflow_results = None
        
        with col2:
            if st.button("✅ Validate Workflow"):
                try:
                    workflow_dict = yaml.safe_load(workflow_yaml)
                    engine = WorkflowEngine()
                    engine.load_workflow_from_dict(workflow_dict)
                    st.success("✅ Workflow is valid!")
                except WorkflowValidationError as e:
                    st.error(f"❌ Validation failed: {str(e)}")
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")

    # Display results outside of columns to use full width
    if st.session_state.workflow_results is not None:
        display_workflow_results(st.session_state.workflow_results)

if __name__ == "__main__":
    main()
